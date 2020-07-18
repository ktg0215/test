from itertools import product
import types
import openpyxl as px
from openpyxl import worksheet
from openpyxl.utils import range_boundaries
import random
import pprint
from scoop import futures
import numpy as np
import pandas as pd
from deap import base
from deap import creator
from deap import algorithms
from deap import tools
from deap import cma
import csv
import pprint
from scoop import futures


wb = px.load_workbook(r'\Users\ktg\Desktop\『目黒店』 シフト表.xlsx', data_only = True)
sheet = wb['原本']


# エクセルからシフトデータ読み込みーーーーーーーーーーーー
# 希望人数
allpa = sheet.cell(row=3, column=3).value
# シフト取得
pp1 = []
def PA2(self,time):
    for i in range(4,19):
        pp1.append(sheet.cell(row=time, column=i).value)
    return pp1
# シフト取得
# 必要人数
k = 4
hope =[]
for aa in range(15):
    hope.append(sheet.cell(row=3, column=k).value)
    k += 1
# hope = [int(i) for i in hope]

bb=4
for j in range(30):
    PA2(1,bb)
    bb= bb + 1
pp = []
for m in range(450):
    if pp1[m] >= 1:
        pp.append(1)
    if pp1[m] == None:
        pp.append(0)
    if pp1[m] == 0:
        pp.append(0)    

name1=sheet.cell(row=4, column=2).value
name2=sheet.cell(row=5, column=2).value
name3=sheet.cell(row=6, column=2).value
name4=sheet.cell(row=7, column=2).value
name5=sheet.cell(row=8, column=2).value
name6=sheet.cell(row=9, column=2).value
name7=sheet.cell(row=10, column=2).value
name8=sheet.cell(row=11, column=2).value
name9=sheet.cell(row=12, column=2).value
name10=sheet.cell(row=13, column=2).value
name11=sheet.cell(row=14, column=2).value
name12=sheet.cell(row=15, column=2).value
name13=sheet.cell(row=16, column=2).value
name14=sheet.cell(row=17, column=2).value
name15=sheet.cell(row=18, column=2).value
name16=sheet.cell(row=19, column=2).value
name17=sheet.cell(row=20, column=2).value
name18=sheet.cell(row=21, column=2).value
name19=sheet.cell(row=22, column=2).value
name20=sheet.cell(row=23, column=2).value
name21=sheet.cell(row=24, column=2).value
name22=sheet.cell(row=25, column=2).value
name23=sheet.cell(row=26, column=2).value
name24=sheet.cell(row=27, column=2).value
name25=sheet.cell(row=28, column=2).value
name26=sheet.cell(row=29, column=2).value
name27=sheet.cell(row=30, column=2).value
name28=sheet.cell(row=31, column=2).value
name29=sheet.cell(row=32, column=2).value
name30=sheet.cell(row=33, column=2).value
tf1=sheet.cell(row=4, column=3).value
tf2=sheet.cell(row=5, column=3).value
tf3=sheet.cell(row=6, column=3).value
tf4=sheet.cell(row=7, column=3).value
tf5=sheet.cell(row=8, column=3).value
tf6=sheet.cell(row=9, column=3).value
tf7=sheet.cell(row=10, column=3).value
tf8=sheet.cell(row=11, column=3).value
tf9=sheet.cell(row=12, column=3).value
tf10=sheet.cell(row=13, column=3).value
tf11=sheet.cell(row=14, column=3).value
tf12=sheet.cell(row=15, column=3).value
tf13=sheet.cell(row=16, column=3).value
tf14=sheet.cell(row=17, column=3).value
tf15=sheet.cell(row=18, column=3).value
tf16=sheet.cell(row=19, column=3).value
tf17=sheet.cell(row=20, column=3).value
tf18=sheet.cell(row=21, column=3).value
tf19=sheet.cell(row=22, column=3).value
tf20=sheet.cell(row=23, column=3).value
tf21=sheet.cell(row=24, column=3).value
tf22=sheet.cell(row=25, column=3).value
tf23=sheet.cell(row=26, column=3).value
tf24=sheet.cell(row=27, column=3).value
tf25=sheet.cell(row=28, column=3).value
tf26=sheet.cell(row=29, column=3).value
tf27=sheet.cell(row=30, column=3).value
tf28=sheet.cell(row=31, column=3).value
tf29=sheet.cell(row=32, column=3).value
tf30=sheet.cell(row=33, column=3).value


# 個人の希望シフト作成ーーーーーーーーーーーーーーーーーーーーーーーーーーー
# pp = [int(i) for i in pp]
t1 = []
i=0
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t1.append("d"+str(g))
    i += 1
    g += 1
t2 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t2.append("d"+str(g))
    i += 1
    g += 1
t3 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t3.append("d"+str(g))
    i += 1
    g += 1
t4 = []
g=1
for ll in range(1,16):
    if '1' == pp[i]:
        t4.append("d"+str(g))
    if 1 == pp[i]:
        t4.append("d"+str(g))
    i += 1
    g += 1
t5 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t5.append("d"+str(g))
    i += 1
    g += 1
t6 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t6.append("d"+str(g))
    i += 1
    g += 1
t7 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t7.append("d"+str(g))
    i += 1
    g += 1
t8 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t8.append("d"+str(g))
    i += 1
    g += 1
t9 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t9.append("d"+str(g))
    i += 1
    g += 1
t10 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t10.append("d"+str(g))
    i += 1
    g += 1
t11 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t11.append("d"+str(g))
    i += 1
    g += 1
t12 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t12.append("d"+str(g))
    i += 1
    g += 1
t13 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t13.append("d"+str(g))
    i += 1
    g += 1
t14 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t14.append("d"+str(g))
    i += 1
    g += 1
t15 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t15.append("d"+str(g))
    i += 1
    g += 1
t16 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t16.append("d"+str(g))
    i += 1
    g += 1
t17 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t17.append("d"+str(g))
    i += 1
    g += 1
t18 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t18.append("d"+str(g))
    i += 1
    g += 1
t19 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t19.append("d"+str(g))
    i += 1
    g += 1
t20 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t20.append("d"+str(g))
    i += 1
    g += 1
t21 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t21.append("d"+str(g))
    i += 1
    g += 1
t22 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t22.append("d"+str(g))
    i += 1
    g += 1
t23 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t23.append("d"+str(g))
    i += 1
    g += 1
t24 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t24.append("d"+str(g))
    i += 1
    g += 1
t25 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t25.append("d"+str(g))
    i += 1
    g += 1
t26 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t26.append("d"+str(g))
    i += 1
    g += 1
t27 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t27.append("d"+str(g))
    i += 1
    g += 1
t28 = []
g=1
for ll in range(1,16):
    if 1 == pp[i]:
        t28.append("d"+str(g))
    i += 1
    g += 1

# 個人の希望シフト作成ーーーーーーーーーーーーーーーーーーーーーーーーーーー
# 従業員を表すクラス
class Employee(object):
  def __init__(self,no,name,manager,wills):
    self.no = no
    self.name = name
    self.manager = manager
    self.wills = wills

  def is_applicated(self, box_name):
    return (box_name in self.wills)

# シフトを表すクラス
# 内部的には15日 * 20人 = 300次元のタプルで構成される
class Shift(object):
  # コマの定義
  SHIFT_BOXES = [
    'd1',
    'd2',
    'd3',
    'd4',
    'd5',
    'd6',
    'd7',
    'd8',
    'd9',
    'd10',
    'd11',
    'd12',
    'd13',
    'd14',
    'd15']

  # 各コマの想定人数
  NEED_PEOPLE = hope

  def __init__(self, list):
    if np.array == None:
      self.make_sample()
    else:
      self.list = list
    self.employees = []

  # ランダムなデータを生成
  def make_sample(self):
    sample_list = np.array([])
    for num in range(allpa*15):
      sample_list.append(random.randint(0, 1))
    self.list = tuple(sample_list)

  # タプルを1ユーザ単位に分割
  def slice(self):
    sliced = []
    start = 0
    for num in range(allpa):
      sliced.append(self.list[start:(start + 15)])
      start = start + 15
    return tuple(sliced)

  # ユーザ別にアサインコマ名を出力する
  def print_inspect(self):
    user_no = 0
    for line in self.slice():
      print("ユーザ%d" % user_no)
      print(line)
      user_no = user_no + 1

      index = 0
      for e in line:
        if e == 1:
          print(self.SHIFT_BOXES[index])
        index = index + 1

  # CSV形式でアサイン結果の出力をする
  def print_csv(self):
    for line in self.slice():
      print(','.join(futures.map(str, line)))

  # TSV形式でアサイン結果の出力をする
  def print_tsv(self):
    for line in self.slice():
      print("\t".join(futures.map(str, line)))


  # ユーザ番号を指定してコマ名を取得する
  def get_boxes_by_user(self, user_no):
    line = self.slice()[user_no]
    return self.line_to_box(line)

  # 1ユーザ分のタプルからコマ名を取得する
  def line_to_box(self, line):
    result = []
    index = 0
    for e in line:
      if e == 1:
        result.append(self.SHIFT_BOXES[index])
      index = index + 1
    return result    

  # コマ番号を指定してアサインされているユーザ番号リストを取得する
  def get_user_nos_by_box_index(self, box_index):
    user_nos = []
    index = 0
    for line in self.slice():
      if line[box_index] == 1:
        user_nos.append(index)
      index += 1
    return user_nos

  # コマ名を指定してアサインされているユーザ番号リストを取得する
  def get_user_nos_by_box_name(self, box_name):
    box_index = self.SHIFT_BOXES.index(box_name)
    return self.get_user_nos_by_box_index(box_index)

  # 想定人数と実際の人数の差分を取得する
  def abs_people_between_need_and_actual(self):
    result = []
    index = 0
    for need in self.NEED_PEOPLE:
      actual = len(self.get_user_nos_by_box_index(index))
      result.append(abs(need - actual))
      index += 1
    return result

  # 応募していないコマにアサインされている件数を取得する
  def not_applicated_assign(self):
    count = 0
    for box_name in self.SHIFT_BOXES:
      user_nos = self.get_user_nos_by_box_name(box_name)
      for user_no in user_nos:
        e = self.employees[user_no]
        if not e.is_applicated(box_name):
          count += 1
    return count

  # true全入れ
  def few_work_user(self):
    result = []
    for user_no in range(allpa):
      e = self.employees[user_no]
      ratio = float(len(self.get_boxes_by_user(user_no))) / float(len(e.wills))
      if e.manager == True:
          if ratio == 1.0:
              result.append(e)
    return result

  # アサインが応募コマ数の50%に満たないユーザを取得
  def few_work_u(self):
    result = []
    for user_no in range(allpa):
      e = self.employees[user_no]
      ratio = float(len(self.get_boxes_by_user(user_no))) / float(len(e.wills))
      if e.manager == False and ratio < 0.7:
        result.append(e)
    return result

# 従業員定義


e0=Employee(0,name1,tf1,t1)
e1=Employee(1,name2,tf2,t2)
e2=Employee(2,name3,tf3,t3)
e3=Employee(3,name4,tf4,t4)
e4=Employee(4,name5,tf5,t5)
e5=Employee(5,name6,tf6,t6)
e6=Employee(6,name7,tf7,t7)
e7=Employee(7,name8,tf8,t8)
e8=Employee(8,name9,tf9,t9)
e9=Employee(9,name10,tf10,t10)
e10=Employee(10,name11,tf11,t11)
e11=Employee(11,name12,tf12,t12)
e12=Employee(12,name13,tf13,t13)
e13=Employee(13,name14,tf14,t14)
e14=Employee(14,name15,tf15,t15)
e15=Employee(15,name16,tf16,t16)
e16=Employee(16,name17,tf17,t17)
e17=Employee(17,name18,tf18,t18)
e18=Employee(18,name19,tf19,t19)
e19=Employee(19,name20,tf20,t20)
e20=Employee(20,name21,tf21,t21)
e21=Employee(21,name22,tf22,t22)
e22=Employee(22,name23,tf23,t23)
e23=Employee(23,name24,tf24,t24)
e24=Employee(24,name25,tf25,t25)
e25=Employee(25,name26,tf26,t26)
e26=Employee(26,name27,tf27,t27)
e27=Employee(27,name28,tf28,t28)

em = [e0,e1,e2,e3,e4,e5,e6,e7,e8,e9,e10,e11,e12,e13,e14,e15,e16,e17,e18,e19,e20,e21,e22,e23,e24,e25,e26,e27]
employees = em[:allpa+1]

creator.create("FitnessPeopleCount", base.Fitness, weights=(-75.0, -300.0,1000.0,-100))
creator.create("Individual", np.ndarray, fitness=creator.FitnessPeopleCount)

toolbox = base.Toolbox()

toolbox.register("attr_bool", random.randint, 0, 1)
toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_bool, allpa*15)
toolbox.register("population", tools.initRepeat, list, toolbox.individual)

def evalShift(individual):
  s = Shift(individual)
  s.employees = employees

  # 想定人数とアサイン人数の差
  people_count_sub_sum = sum(s.abs_people_between_need_and_actual()) / allpa*15
  # 応募していない時間帯へのアサイン数
  not_applicated_count = s.not_applicated_assign() / allpa*15
  # アサイン数が応募数の半分以下の従業員数
  few_work_user = len(s.few_work_user()) / allpa 
  # アサイン数が応募数の半分以下の従業員数
  few_work_u = len(s.few_work_u()) / allpa
  return (not_applicated_count, people_count_sub_sum,few_work_user,few_work_u)

toolbox.register("evaluate", evalShift)
# 交叉関数を定義(二点交叉)
toolbox.register("mate", tools.cxTwoPoint)

# 変異関数を定義(ビット反転、変異隔離が5%ということ?)
toolbox.register("mutate", tools.mutFlipBit, indpb=0.05)

# 選択関数を定義(トーナメント選択、tournsizeはトーナメントの数？)
toolbox.register("select", tools.selTournament, tournsize=3)
def cxTwoPointCopy(ind1, ind2):
    size = len(ind1)
    cxpoint1 = random.randint(1, size)
    cxpoint2 = random.randint(1, size - 1)
    if cxpoint2 >= cxpoint1:
        cxpoint2 += 1
    else: # Swap the two cx points
        cxpoint1, cxpoint2 = cxpoint2, cxpoint1

    ind1[cxpoint1:cxpoint2], ind2[cxpoint1:cxpoint2] = ind2[cxpoint1:cxpoint2].copy(), ind1[cxpoint1:cxpoint2].copy()
        
    return ind1, ind2
    
    
toolbox.register("evaluate", evalShift)
toolbox.register("mate", cxTwoPointCopy)
toolbox.register("mutate", tools.mutFlipBit, indpb=0.05)
toolbox.register("select", tools.selTournament, tournsize=3)

if __name__ == "__main__":
        
    pop = toolbox.population(n=300)
    
    hof = tools.HallOfFame(1, similar=np.array_equal)
    
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register("avg", np.mean)
    stats.register("std", np.std)
    stats.register("min", np.min)
    stats.register("max", np.max)
    
    algorithms.eaSimple(pop, toolbox, cxpb=0.5, mutpb=0.6, ngen=10, stats=stats,halloffame=hof)
    
    print("-- 進化終了 --")
    best_ind = tools.selBest(pop, 1)[0]
    s = Shift(best_ind)
        
    # 必要人数
need = Shift.NEED_PEOPLE
need.insert(0,"必要人数")
day = Shift.SHIFT_BOXES
day.insert(0,"日")
aa1 = []
for aa in range(allpa*15):
        if best_ind[aa] == 1:
            aa1.append(pp1[aa])
        if best_ind[aa] == 0:
            aa1.append(None)     
pl = []
start = 0
for num in range(allpa):
    pl.append(aa1[start:(start + 15)])
    start = start + 15
i=0
p1=[]
for nn in range(allpa):
    p1.append(employees[i].name)
    p1[len(p1):len(p1)] = pl[i]
    i+=1
pa= []
pa.append(day)
pa.append(need)
start = 0
for num in range(allpa):
    pa.append(p1[start:(start + 16)])
    start = start + 16
df=pd.DataFrame(pa)
a =['d1','d2','d3','d4','d5','d6','d7','d8','d9','d10','d11','d12','d13','d14','d15',
]
pd.DataFrame(columns=a)
df.replace(None,0)
# 必要人数
need = Shift.NEED_PEOPLE
df = pd.DataFrame(pa)
df.replace(0,None)
with pd.ExcelWriter(r'\Users\ktg\Desktop\『目黒店』 シフト表.xlsx',
                    engine="openpyxl", mode="a") as ew:
    df.to_excel(ew, sheet_name="貼り付け")

    
    

   